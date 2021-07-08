import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl
import time


# Finds the crunchbase page from google search results
def FindCrunchbasePage(search_result):
    for link in search_result.findAll('a'):
        if "Crunchbase" in link.text:
            crunchbase_page = link['href']
            print(f'Crunchbase URL: {crunchbase_page}')

            return crunchbase_page


def FindVCSite(crunch_soup):
    try:
        for elem in crunch_soup.findAll('a', {'class': 'component--field-formatter link-accent ng-star-inserted'}):
            company_site = elem['href']
            return company_site
    except:
        print('no url!')
        return ''


def FindVCEmail(crunch_soup):
    span_list = []

    try:
        for elem in crunch_soup.findAll('span', {'class': 'ng-star-inserted'}):
            span_list.append(elem.text) 
    except:
        print('no text found!')

    # Find the "Contact Email" among the list of span elements
    try:
        email_heading_idx = span_list.index('Contact Email\xa0')
        company_email = span_list[email_heading_idx + 1]

        return company_email
    except:
        return ''


# Finds company website and email
def FindCompanyInfo(crunch_url):
    driver.get(crunch_url)
    crunchbase_soup = BeautifulSoup(driver.page_source, 'html.parser')

    website = FindVCSite(crunchbase_soup)
    email = FindVCEmail(crunchbase_soup)

    return [website, email]


def PerformGoogleSearch(vc_name):
    URL = 'https://www.google.ca/search?q='
    raw_company_name = vc_name
    parameters = raw_company_name.replace(' ', '+')

    # Make a get request using selenium
    driver.get(URL + parameters)

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    return soup


def ExportToExcel(excel_file, website, contact, row):

    website_cell = ws['B' + str(row)]
    contact_cell = ws['C' + str(row)]

    website_cell.value = website
    contact_cell.value = contact


if __name__ == "__main__":
    driver = webdriver.Firefox()

    output_file = 'vc_info.xlsx'

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    for row in range(110, 455):
        vc = ws['A' + str(row + 1)].value

        google_soup = PerformGoogleSearch(vc)

        cbase_page = FindCrunchbasePage(google_soup)

        if cbase_page is not None:
            info = FindCompanyInfo(cbase_page)
            print(f'Company site: {info[0]}')
            print(f'Company email: {info[1]}')
            ExportToExcel(output_file, info[0], info[1], row + 1)

        else:
            print(f'No crunchbase profile found for {vc}!')

        time.sleep(1)

    wb.save(output_file)

    driver.close()
    


